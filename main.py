from openpyxl import load_workbook
import requests
import sender
import gi
gi.require_version("Gtk", "3.0")
from gi.repository import Gtk

class PersonPayslip:

    def __init__(self):
        self.name = ""
        self.email = ""
        self.pdf_location = ""

    def __str__(self):
        return self.name + " - " + self.name + ", email: " + self.email + " AT " + self.pdf_location


class MyWindow(Gtk.Window):

    def __init__(self, persons):
        super().__init__(title="Mini Email Sender")
        self.persons = persons
        self.checkboxes = []

        box_v = Gtk.Box(orientation="vertical", spacing=0, margin=20)
        self.add(box_v)
        
        grid = Gtk.Grid(row_spacing=10, column_spacing=10, margin_right=10)

        # Show persons list
        row = 0
        for pers in self.persons:
            checkbox = Gtk.CheckButton()
            self.checkboxes.append( checkbox )
            name_label = Gtk.Label( label=pers.name, halign=Gtk.Align.START )
            email_label = Gtk.Label( label=pers.email, halign=Gtk.Align.START )
            grid.attach( checkbox, 0, row, 1, 1)
            grid.attach( name_label, 1, row, 1 , 1)
            grid.attach( email_label, 2, row, 1 , 1)
            row += 1
        
        box_v.pack_start(grid, True, True, 0)

        # Butonul de Generare
        self.button = Gtk.Button(label="Send", margin=10, margin_top=40)
        self.button.connect("clicked", self.on_send)
        box_v.pack_start(self.button, True, True, 0)

    def on_send(self, widget):
        to_send = []
        for index, checkbox in enumerate(self.checkboxes):
            if checkbox.get_active() == True:
                to_send.append( self.persons[index] )

        for payslip in to_send:
            print("Sending to " + p.email )
            sender.send_email(payslip.email, payslip.pdf_location)


wb = load_workbook(filename = 'payslips.xlsx')
ws = wb.active

payslips = []
for index, row in enumerate(ws.iter_rows(min_row=2, max_row=9, min_col=1, max_col=3)):
    if index < 1:
        continue
    ps = PersonPayslip()
    ps.name = ws.cell(row=index+1, column=1).value
    ps.email = ws.cell(row=index+1, column=2).value
    ps.pdf_location = ws.cell(row=index+1, column=3).value
    payslips.append(ps)

for i, p in enumerate(payslips):
    print("[%d] %s" % (i+1, p))


gui = MyWindow(payslips)
gui.connect( "destroy", Gtk.main_quit )
gui.show_all()
Gtk.main()

